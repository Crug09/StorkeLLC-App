import os
import sys
import datetime
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.lang import Builder

log_file = 'Log.txt'
if os.path.exists(log_file):
    pass
else:
    f = open(log_file, 'x')
def log(x):
    f = open(log_file, 'a')
    f.write(x)
    f.close()

log('Starting App\n')

class Main_Screen(Screen):
    pass


class Megger_Test_Screen(Screen):
    # log('Opened Megger_Test_Screen\n')
    def process_input(self):
        log('Starting megger test input\n')

        customer_name = self.ids.customer_name.text
        customer_name = customer_name.replace(' ', '_')
        customer_name = customer_name.title()
        log(f'Customers name: {customer_name}\n')
        if customer_name:
            #Adds the data
            phase_labels = ['Phase 1', 'Phase 2', 'Phase 3', ' ']
            underground_wire = self.ids.underground_wire.active
            tower_disconnect_switch =self.ids.tower_disconnect_switch.active
            tower_wire = self.ids.tower_wire.active
            junction_box = self.ids.junction_box.active
            slip_rings = self.ids.slip_rings.active
            brush_block = self.ids.brush_block.active
            altenator = self.ids.altenator.active

            phase_1 = self.ids.phase_1.text
            phase_2 = self.ids.phase_2.text
            phase_3 = self.ids.phase_3.text


            if phase_1 and phase_2 and phase_3:

                #Adds all of the variables of the test
                megger_test = []
                if underground_wire:
                    megger_test.append('Underground_Wire')
                if tower_disconnect_switch:
                    megger_test.append('Tower_Disconnect_Switch')
                if tower_wire:
                    megger_test.append('Tower_Wire')
                if junction_box:
                    megger_test.append('Junction_Box')
                if slip_rings:
                    megger_test.append('Slip_Rings')
                if brush_block:
                    megger_test.append('Brush_Block')
                if altenator:
                    megger_test.append('Altenator')

                megger_test = '__'.join(megger_test)
                log('Has all the data for megger test input\n')

                if megger_test != '':
                    log(f'Phase 1: {phase_1}, Phase 2: {phase_2}, Phase 3: {phase_3}\n')
                    log(f'Megger Test Variables: {megger_test}\n')
                    #Makes a file name with customer name and the days date
                    today_date = datetime.datetime.now().strftime('%Y-%m-%d')
                    file_name = f"{customer_name}_Megger_Test_{today_date}.xlsx"

                    phase_results = [int(phase_1), int(phase_2), int(phase_3), ' ']
                    
                    #Sees if there is already that files name
                    if os.path.exists(file_name):
                        workbook = load_workbook(file_name)
                        sheet = workbook.active
                        workbook.save(file_name)
                        log('Loaded previously made megger test file\n')
                    else:
                        workbook = Workbook()
                        sheet = workbook.active
                        sheet.title = "Sheet1"
                        workbook.save(file_name)
                        log('Made new megger test file\n')

                    workbook = load_workbook(file_name)

                    sheet = workbook.active
                    row_number = 1
                    row = sheet[row_number]

                    #gets data from the cells in the first row to see if test is in it
                    components = [cell.value for cell in row if cell.value is not None]

                    #adds new column if the compnent list isnt in it
                    if megger_test in components:
                        pass
                        log('Did not make a new column')                        
                    else:
                        last_column = sheet.max_column
                        new_column = last_column + 1
                        cell = sheet.cell(row=row_number, column=new_column)
                        cell.value = megger_test
                        workbook.save(file_name) 
                        log('Made a new column\n')


                    #color of data valadation
                    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red background
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
                    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green background

                    #Find the list of compents that was inputted
                    target_column = None
                    for cell in sheet[1]: 
                        if cell.value == megger_test:
                            target_column = cell.column
                            break
                    if target_column is None:
                        self.ids.megger_test_id.text = 'Error has occured'
                        log('Error has occured with finding the variables\n')
                    else:
                        start_row = 1
                        while sheet.cell(row=start_row, column=target_column).value is not None:
                            start_row += 1
                        log(f'Starting putting in data into {customer_name}\n')
                        #Adds data
                        for i, (item, phase) in enumerate(zip(phase_results, phase_labels)):
                            phase_cell = sheet.cell(row=start_row + i, column=1)
                            phase_cell.value = phase
                            log(f'Inputing {phase} into row {start_row+i}, column: 1\n')
                            
                            data_cell = sheet.cell(row=start_row + i, column=target_column)
                            data_cell.value = item
                            log(f'Inputing {item} into row {start_row+i}, column: {target_column}\n')

                        for row in range(start_row, start_row + len(phase_results)): 
                            cell = sheet.cell(row=row, column=target_column)
                            
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                if 0 <= cell.value <= 349:
                                    cell.fill = red_fill
                                elif 350 <= cell.value <= 524:
                                    cell.fill = yellow_fill
                                elif 525 <= cell.value <= 550:
                                    cell.fill = green_fill
                                log('Doing color formatting\n')

                        workbook.save(file_name)
                        log('Saving megger test file\n')
                        megger_test = megger_test.replace('__', ' and ')
                        megger_test = megger_test.replace('_', ' ')
                        self.ids.megger_test_id.text = f'Added:\n{megger_test}\nWith the data being:{phase_results[:3]}'
                else:
                    self.ids.megger_test_id.text = 'Enter a component'
                    log('Does not have a compent marked\n')
            else:
                self.ids.megger_test_id.text = 'Enter all phases'
                log('Does not have all three phases listed\n')
        else:
            self.ids.megger_test_id.text = 'Enter name of customer' 
            log('Does not have the customers name written down\n') 

    def clear_megger_spec(self, instance):
        x = instance.text
        if x == 'X':
            self.ids.customer_name.text = ''
            log('Removed text for customer name\n')
        if x == 'X ':
            self.ids.underground_wire.active = False
            self.ids.tower_disconnect_switch.active = False
            self.ids.tower_wire.active = False
            self.ids.junction_box.active = False
            self.ids.slip_rings.active = False
            self.ids.brush_block.active = False
            self.ids.altenator.active = False
            log('Removed variables for megger test\n')
        if x == ' X ':
            self.ids.phase_1.text = ''
            log('Removed text for phase 1\n')
        if x == ' X  ':
            self.ids.phase_2.text = ''
            log('Removed text for phase 2\n')
        if x == '  X  ':
            self.ids.phase_3.text = ''
            log('Removed text for phase 3\n')


    def clear_inputs(self):
        self.ids.customer_name.text = ''
        self.ids.underground_wire.active = False
        self.ids.tower_disconnect_switch.active = False
        self.ids.tower_wire.active = False
        self.ids.junction_box.active = False
        self.ids.slip_rings.active = False
        self.ids.brush_block.active = False
        self.ids.altenator.active = False
        self.ids.phase_1.text = ''
        self.ids.phase_2.text = ''
        self.ids.phase_3.text = ''
        self.ids.megger_test_id.text = ''
        log('Removed all variables from Megger_Test_Screen\n')

# class Names_Screen(Screen):
#     def update_file_list(self, file_names):
#         if file_names:
#             self.ids.file_list_label.text = '\n'.join(file_names)
#         else:
#             self.ids.file_list_label.text = "No previous test files found."


class Distance_Calc_Screen(Screen):
    # log('Opened Distance_Calc_Screen\n')
    def process_distance(self):
        log('Starting distance calculation\n')
        #Adds data
        file_name = self.ids.name.text
        customers = self.ids.customers.text
        hours = self.ids.hours.text
        rate = self.ids.rate.text
        customers = customers.split(',')
        hours = hours.split(',')
        log(f'Customers: {customers}, Hours: {hours}, Rate: {rate}\n')
        if len(customers) == len(hours) and customers[0] != '':
            hours = [float(num.strip()) for num in hours]
            total_hours = 0
            for x in range(len(hours)):
                total_hours = float(hours[x]) + float(total_hours)
            log(f'Total Hours: {total_hours}\n')
            max_pay = float(max(hours)) * float(rate)
            log(f'Maximun paying amount is {max_pay} with the highest hours traveled at {max(hours)}\n')
            customers_pay = []
            for x in range(len(hours)):
                customers_pay.append(((float(hours[x]) * float(max_pay)) / float(total_hours)) * 2)
            log(f'Customers pay: {customers_pay}\n')

            #Sees if there is a file with that name
            file_name = f'{file_name}.xlsx'
            if os.path.exists(file_name):
                workbook = load_workbook(file_name)
                sheet = workbook.active
                workbook.save(file_name)
                log('Loaded previously made distance calculation file\n')
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                workbook.save(file_name)
                log('Made new distance calculation file\n')

            #Removes all old data and puts new data
            workbook = load_workbook(file_name)
            sheet = workbook.active
            sheet.delete_cols(idx=1, amount=3)
            log('Deleted columns 1, 2 and 3 data\n')
            sheet['A1'] = 'Name'
            sheet['B1'] = 'Hours'
            sheet['C1'] = 'Payment'
            workbook.save(file_name)
            for x in range(len(hours)):
                sheet[f'A{x+2}'] = customers[x]
                log(f'Inputed Customer: {customers[x]} into A{x+2}\n')
                sheet[f'B{x+2}'] = hours[x]
                log(f'Inputed {hours[x]} hours into B{x+2}\n')
                sheet[f'C{x+2}'] = f'${round(customers_pay[x], 2)}'
                log(f'Inputed ${round(customers_pay[x], 2)} into C{x+2}\n')
            
            workbook.save(file_name)
            self.ids.distance_calc_id.text = f'File: {file_name} has been saved'
            log('Saved File\n')

        elif rate == '':
            self.ids.distance_calc_id.text = 'You have not enough data'
            log('There is no rate\n')
        else:
            self.ids.distance_calc_id.text = 'There are not the same amount of customers and hours'
            log('There is not the same amount of hours and customers\n')

    def clear_distance_spec(self, instance):
        x = instance.text
        if x == 'X ':
            self.ids.name.text = ''
            log('Removed text for file name\n')
        if x == ' X ':
            self.ids.customers.text = ''
            log('Removed text for customer names\n')
        if x == ' X  ':
            self.ids.hours.text = ''
            log('Removed text for the hours\n')
        if x == '  X  ':
            self.ids.rate.text = ''
            log('Removed text for the rate\n')

    def clear_distance(self, instance):
        self.ids.distance_calc_id.text = ''
        self.ids.name.text = ''
        self.ids.customers.text = ''
        self.ids.hours.text = ''
        self.ids.rate.text = ''
        log('Deleted inputs for Distance_Calc_Screen\n')


class ScreenManagement(ScreenManager):
    pass

class MyKivyApp(App):
    def build(self):            
        if hasattr(sys, '_MEIPASS'):
            kv_file1_path = os.path.join(sys._MEIPASS, 'main.kv')
            kv_file2_path = os.path.join(sys._MEIPASS, 'distance_calc.kv')
            kv_file3_path = os.path.join(sys._MEIPASS, 'megger_test.kv')
        else:
            # If not bundled, use the regular paths
            kv_file1_path = 'main.kv'
            kv_file2_path = 'distance_calc.kv'
            kv_file3_path = 'megger_test.kv'
        Builder.load_file(kv_file1_path)
        Builder.load_file(kv_file2_path)
        Builder.load_file(kv_file3_path)
        
        sm = ScreenManagement()
        sm.add_widget(Main_Screen(name='main_screen'))
        sm.add_widget(Distance_Calc_Screen(name='distance_calc_screen'))
        sm.add_widget(Megger_Test_Screen(name='megger_test_screen'))
        # sm.add_widget(Names_Screen(name='name_screen'))
        return sm

    
    # def show_excel_files(self):
    #     directory_path = '.'  
    #     file_names = [f for f in os.listdir(directory_path) if f.endswith('.xlsx')]
    #     customer_names = []
    #     x = 0
    #     for i in file_names:
    #         base_name = os.path.splitext(file_names[x])[0]
    #         parts = re.split('[._]', base_name)
    #         customer_names.append(parts[0] + " " + parts[1] + " " + parts[4] )
    #         x+=1
    #     name_screen = self.root.get_screen('name_screen')
    #     name_screen.update_file_list(customer_names)
    #     self.root.current = 'name_screen'
        

if __name__ == '__main__':
    MyKivyApp().run()
